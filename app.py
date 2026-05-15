"""
Sistema de Reservas de Espacios — Sprint 3
Módulo de autenticación + gestión de perfil + frontend por rol
"""

from asyncio import coroutines
from asyncio import coroutines
from flask import Flask, render_template, request, redirect, url_for, session, flash, make_response, jsonify
from flask_mail import Mail, Message
import bcrypt
import traceback
from database import (crear_base_de_datos, Session as DBSession, Usuario, Admin,
                       Docente, Administrativo, PersonaExterna, CONFLICTOS,
                       PersonaExternaReserva)
from datetime import date, timedelta, datetime, time as dtime
from itsdangerous import URLSafeTimedSerializer
import os

app = Flask(__name__)
app.secret_key = "clave_secreta_reservas_2025"
app.permanent_session_lifetime = timedelta(minutes=15)

# ── Configuración de correo ──────────────────────────────────────────────────
app.config["MAIL_SERVER"]   = "smtp.gmail.com"
app.config["MAIL_PORT"]     = 587
app.config["MAIL_USE_TLS"]  = True
app.config["MAIL_USERNAME"] = "reservas.poligranco@gmail.com"
app.config["MAIL_PASSWORD"] = "hzro gypt dspm sbvc"
app.config["MAIL_DEFAULT_SENDER"] = ("Sistema de Reservas — Poli", "reservas.poligranco@gmail.com")

mail = Mail(app)
serializer = URLSafeTimedSerializer(app.secret_key)

@app.after_request
def add_header(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response
# ── Funciones de correo ──────────────────────────────────────────────────────

def enviar_correo_reserva_creada(usuario, reserva, espacio):
    """Notifica al usuario que su reserva fue recibida."""
    try:
        msg = Message(
            subject="✅ Solicitud de reserva recibida — Politécnico Grancolombiano",
            recipients=[usuario.correo]
        )
        msg.html = f"""
        <div style="font-family:Arial,sans-serif; max-width:600px; margin:0 auto;">
          <div style="background:#0d2340; padding:20px; text-align:center;">
            <h2 style="color:#00b4d8; margin:0;">Sistema de Reservas</h2>
            <p style="color:rgba(255,255,255,.6); margin:4px 0 0;">Politécnico Grancolombiano</p>
          </div>
          <div style="padding:30px; background:#f4f6f9;">
            <p>Hola <strong>{usuario.nombre}</strong>,</p>
            <p>Tu solicitud de reserva ha sido recibida y está <strong>en revisión</strong>. Te notificaremos cuando sea aprobada.</p>
            <div style="background:white; border-radius:8px; padding:20px; border-left:4px solid #00b4d8; margin:20px 0;">
              <h3 style="color:#0d2340; margin-top:0;">Detalles de la reserva</h3>
              <p><strong>Espacio:</strong> {espacio.nombre}</p>
              <p><strong>Evento:</strong> {reserva.tipo_evento}</p>
              <p><strong>Fecha:</strong> {reserva.fecha.strftime('%d/%m/%Y')}</p>
              <p><strong>Horario:</strong> {reserva.hora_inicio.strftime('%H:%M')} – {reserva.hora_fin.strftime('%H:%M')}</p>
              <p><strong>Personas:</strong> {reserva.cant_personas}</p>
              <p><strong>Estado:</strong> <span style="color:#e6a817;">⏳ En revisión</span></p>
            </div>
            <p style="color:#6b7c93; font-size:13px;">Este correo fue generado automáticamente. Por favor no respondas a este mensaje.</p>
          </div>
          <div style="background:#0d2340; padding:12px; text-align:center;">
            <p style="color:rgba(255,255,255,.4); font-size:12px; margin:0;">Politécnico Grancolombiano — Institución Universitaria</p>
          </div>
        </div>
        """
        try:
            mail.send(msg)
        except:
            pass
    except Exception as e:
        print(f"Error enviando correo de reserva creada: {e}")


def enviar_correo_reserva_aprobada(usuario, reserva, espacio):
    """Notifica al usuario que su reserva fue aprobada."""
    try:
        msg = Message(
            subject="🎉 Reserva confirmada — Politécnico Grancolombiano",
            recipients=[usuario.correo]
        )
        msg.html = f"""
        <div style="font-family:Arial,sans-serif; max-width:600px; margin:0 auto;">
          <div style="background:#0d2340; padding:20px; text-align:center;">
            <h2 style="color:#00b4d8; margin:0;">Sistema de Reservas</h2>
            <p style="color:rgba(255,255,255,.6); margin:4px 0 0;">Politécnico Grancolombiano</p>
          </div>
          <div style="padding:30px; background:#f4f6f9;">
            <p>Hola <strong>{usuario.nombre}</strong>,</p>
            <p>🎉 Tu reserva ha sido <strong>confirmada</strong>. A continuación los detalles:</p>
            <div style="background:white; border-radius:8px; padding:20px; border-left:4px solid #28a745; margin:20px 0;">
              <h3 style="color:#0d2340; margin-top:0;">Detalles de la reserva</h3>
              <p><strong>Espacio:</strong> {espacio.nombre}</p>
              <p><strong>Evento:</strong> {reserva.tipo_evento}</p>
              <p><strong>Fecha:</strong> {reserva.fecha.strftime('%d/%m/%Y')}</p>
              <p><strong>Horario:</strong> {reserva.hora_inicio.strftime('%H:%M')} – {reserva.hora_fin.strftime('%H:%M')}</p>
              <p><strong>Personas:</strong> {reserva.cant_personas}</p>
              <p><strong>Estado:</strong> <span style="color:#28a745;">✅ Confirmada</span></p>
            </div>
            <p style="color:#6b7c93; font-size:13px;">Este correo fue generado automáticamente. Por favor no respondas a este mensaje.</p>
          </div>
          <div style="background:#0d2340; padding:12px; text-align:center;">
            <p style="color:rgba(255,255,255,.4); font-size:12px; margin:0;">Politécnico Grancolombiano — Institución Universitaria</p>
          </div>
        </div>
        """
        try:
            mail.send(msg)
        except:
            pass
    except Exception as e:
        print(f"Error enviando correo de reserva aprobada: {e}")


def enviar_correo_reserva_rechazada(usuario, reserva, espacio, justificacion=None):
    """Notifica al usuario que su reserva fue rechazada, incluyendo la razón del rechazo."""
    try:
        razon_html = ""
        if justificacion:
            razon_html = f"""
              <div style="background:#fff3f3; border:1px solid #dc3545; border-radius:8px; padding:16px; margin:20px 0;">
                <h4 style="color:#dc3545; margin:0 0 8px; font-size:14px;">📋 Razón del rechazo</h4>
                <p style="margin:0; color:#333; font-size:14px; line-height:1.6;">{justificacion}</p>
              </div>"""

        msg = Message(
            subject="❌ Reserva rechazada — Politécnico Grancolombiano",
            recipients=[usuario.correo]
        )
        msg.html = f"""
        <div style="font-family:Arial,sans-serif; max-width:600px; margin:0 auto;">
          <div style="background:#0d2340; padding:20px; text-align:center;">
            <h2 style="color:#00b4d8; margin:0;">Sistema de Reservas</h2>
            <p style="color:rgba(255,255,255,.6); margin:4px 0 0;">Politécnico Grancolombiano</p>
          </div>
          <div style="padding:30px; background:#f4f6f9;">
            <p>Hola <strong>{usuario.nombre}</strong>,</p>
            <p>Lamentamos informarte que tu solicitud de reserva ha sido <strong>rechazada</strong> por el administrador.</p>
            <div style="background:white; border-radius:8px; padding:20px; border-left:4px solid #dc3545; margin:20px 0;">
              <h3 style="color:#0d2340; margin-top:0;">Detalles de la reserva</h3>
              <p><strong>Espacio:</strong> {espacio.nombre}</p>
              <p><strong>Evento:</strong> {reserva.tipo_evento}</p>
              <p><strong>Fecha:</strong> {reserva.fecha.strftime('%d/%m/%Y')}</p>
              <p><strong>Horario:</strong> {reserva.hora_inicio.strftime('%H:%M')} – {reserva.hora_fin.strftime('%H:%M')}</p>
              <p><strong>Personas:</strong> {reserva.cant_personas}</p>
              <p><strong>Estado:</strong> <span style="color:#dc3545;">❌ Rechazada</span></p>
            </div>
            {razon_html}
            <p>Si tienes dudas, comunícate con la administración del Politécnico.</p>
            <p style="color:#6b7c93; font-size:13px;">Este correo fue generado automáticamente. Por favor no respondas a este mensaje.</p>
          </div>
          <div style="background:#0d2340; padding:12px; text-align:center;">
            <p style="color:rgba(255,255,255,.4); font-size:12px; margin:0;">Politécnico Grancolombiano — Institución Universitaria</p>
          </div>
        </div>
        """
        try:
            mail.send(msg)
        except:
            pass
    except Exception as e:
        print(f"Error enviando correo de reserva rechazada: {e}")


def enviar_correo_reserva_cancelada(usuario, reserva, espacio):
    """Notifica al usuario que su reserva fue cancelada por él mismo."""
    try:
        msg = Message(
            subject="🚫 Reserva cancelada — Politécnico Grancolombiano",
            recipients=[usuario.correo]
        )
        msg.html = f"""
        <div style="font-family:Arial,sans-serif; max-width:600px; margin:0 auto;">
          <div style="background:#0d2340; padding:20px; text-align:center;">
            <h2 style="color:#00b4d8; margin:0;">Sistema de Reservas</h2>
            <p style="color:rgba(255,255,255,.6); margin:4px 0 0;">Politécnico Grancolombiano</p>
          </div>
          <div style="padding:30px; background:#f4f6f9;">
            <p>Hola <strong>{usuario.nombre}</strong>,</p>
            <p>Tu reserva ha sido <strong>cancelada</strong> exitosamente según tu solicitud.</p>
            <div style="background:white; border-radius:8px; padding:20px; border-left:4px solid #6c757d; margin:20px 0;">
              <h3 style="color:#0d2340; margin-top:0;">Detalles de la reserva cancelada</h3>
              <p><strong>Espacio:</strong> {espacio.nombre}</p>
              <p><strong>Evento:</strong> {reserva.tipo_evento}</p>
              <p><strong>Fecha:</strong> {reserva.fecha.strftime('%d/%m/%Y')}</p>
              <p><strong>Horario:</strong> {reserva.hora_inicio.strftime('%H:%M')} – {reserva.hora_fin.strftime('%H:%M')}</p>
              <p><strong>Personas:</strong> {reserva.cant_personas}</p>
              <p><strong>Estado:</strong> <span style="color:#6c757d;">🚫 Cancelada</span></p>
            </div>
            <p>Si necesitas reservar un espacio nuevamente, puedes hacerlo desde el sistema de reservas.</p>
            <p style="color:#6b7c93; font-size:13px;">Este correo fue generado automáticamente. Por favor no respondas a este mensaje.</p>
          </div>
          <div style="background:#0d2340; padding:12px; text-align:center;">
            <p style="color:rgba(255,255,255,.4); font-size:12px; margin:0;">Politécnico Grancolombiano — Institución Universitaria</p>
          </div>
        </div>
        """
        try:
            mail.send(msg)
        except:
            pass
    except Exception as e:
        print(f"Error enviando correo de cancelación: {e}")


def enviar_correo_recuperacion(correo, nombre, token):
    """Envía el correo de recuperación de contraseña."""
    try:
        link = url_for("recuperar_contrasena_reset", token=token, _external=True)
        msg = Message(
            subject="🔐 Recuperación de contraseña — Politécnico Grancolombiano",
            recipients=[correo]
        )
        msg.html = f"""
        <div style="font-family:Arial,sans-serif; max-width:600px; margin:0 auto;">
          <div style="background:#0d2340; padding:20px; text-align:center;">
            <h2 style="color:#00b4d8; margin:0;">Sistema de Reservas</h2>
            <p style="color:rgba(255,255,255,.6); margin:4px 0 0;">Politécnico Grancolombiano</p>
          </div>
          <div style="padding:30px; background:#f4f6f9;">
            <p>Hola <strong>{nombre}</strong>,</p>
            <p>Recibimos una solicitud para restablecer la contraseña de tu cuenta.</p>
            <div style="text-align:center; margin:30px 0;">
              <a href="{link}" style="background:#00b4d8; color:#0d2340; padding:14px 28px; border-radius:8px; text-decoration:none; font-weight:bold; font-size:16px;">
                Restablecer contraseña
              </a>
            </div>
            <p style="color:#6b7c93; font-size:13px;">Este enlace expira en <strong>30 minutos</strong>. Si no solicitaste este cambio, ignora este correo.</p>
          </div>
          <div style="background:#0d2340; padding:12px; text-align:center;">
            <p style="color:rgba(255,255,255,.4); font-size:12px; margin:0;">Politécnico Grancolombiano — Institución Universitaria</p>
          </div>
        </div>
        """
        try:
            mail.send(msg)
        except:
            pass
        return True
    except Exception as e:
        print(f"Error enviando correo de recuperación: {e}")
        return False

crear_base_de_datos()

db = DBSession()

@app.teardown_appcontext
def shutdown_session(exception=None):
    DBSession.remove()

@app.after_request
def agregar_headers_no_cache(response):
    """Evita que el navegador guarde páginas en caché."""
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0, post-check=0, pre-check=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "-1"
    return response


@app.before_request
def limpiar_sesion_al_iniciar():
    """Limpia la sesión y carga espacios predefinidos la primera vez que arranca el servidor."""
    if not app.config.get("SESION_LIMPIADA"):
        session.clear()
        app.config["SESION_LIMPIADA"] = True
        _cargar_espacios_predefinidos()

def _cargar_espacios_predefinidos():
    """Crea los espacios del auditorio si no existen aún."""
    from database import Espacio
    # Espacios pequeños: $100.000 COP para externos
    # Auditorio Completo: $150.000 COP para externos
    # Institucionales y docentes: sin costo
    espacios_default = [
        {"nombre": "Sala B1",               "capacidad": 50,  "costo_externo": 100000},
        {"nombre": "Sala B2",               "capacidad": 50,  "costo_externo": 100000},
        {"nombre": "Sala B3",               "capacidad": 100, "costo_externo": 100000},
        {"nombre": "Sala de Música y Baile", "capacidad": 100, "costo_externo": 100000},
        {"nombre": "Auditorio Completo",    "capacidad": 200, "costo_externo": 150000},
    ]
    for e in espacios_default:
        existe = db.query(Espacio).filter_by(nombre=e["nombre"]).first()
        if not existe:
            db.add(Espacio(nombre=e["nombre"], capacidad=e["capacidad"],
                          costo_externo=e["costo_externo"]))
    db.commit()

# ─── Helpers ───────────────────────────────────────────────────────────────

def reserva_ha_finalizado(reserva):
    """Retorna True si la fecha+hora_fin de la reserva ya pasó completamente."""
    return datetime.now() > datetime.combine(reserva.fecha, reserva.hora_fin)

def usuario_logueado():
    """Retorna el objeto Usuario si hay sesión activa, o None."""
    uid = session.get("usuario_id")
    if uid:
        return db.get(Usuario, uid)
    return None

def requiere_login(f):
    """Decorador: redirige al login si no hay sesión."""
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        u = usuario_logueado()
        if not u:
            flash("Debes iniciar sesión primero.", "warning")
            return redirect(url_for("login"))
        if not u.activo:
            session.clear()
            flash("Tu cuenta ha sido desactivada. Contacta al administrador.", "danger")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapper

def requiere_admin(f):
    """Decorador: solo permite acceso a admins."""
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        u = usuario_logueado()
        if not u or u.rol != "admin":
            flash("No tienes permiso para acceder a esa página.", "danger")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return wrapper


# ─── Rutas públicas ────────────────────────────────────────────────────────

@app.route("/")
def index():
    if usuario_logueado():
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if usuario_logueado():
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        correo     = request.form.get("correo", "").strip()
        contrasena = request.form.get("contrasena", "").strip()

        # Flujo alternativo 2a: campos vacíos — validar ANTES de conectar a la BD
        if not correo or not contrasena:
            flash("Todos los campos son obligatorios.", "danger")
            return render_template("login.html")

        # Flujo alternativo 3a: credenciales incorrectas
        usuario = db.query(Usuario).filter_by(correo=correo).first()
        def verificar_pw(plain, stored):
            try:
                return bcrypt.checkpw(plain.encode(), stored.encode())
            except Exception:
                return plain == stored
        if usuario and verificar_pw(contrasena, usuario.contrasena):
            # Verificar si la cuenta está activa
            if not usuario.activo:
                flash("Tu cuenta ha sido desactivada. Contacta al administrador.", "danger")
                return render_template("login.html")
            session.permanent = True
            session["usuario_id"] = usuario.id
            flash(f"Bienvenido, {usuario.nombre}.", "success")
            return redirect(url_for("dashboard"))
        else:
            flash("Correo o contraseña incorrectos.", "danger")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("Sesi\u00f3n cerrada correctamente.", "info")
    resp = make_response(redirect(url_for("login")))
    # Forzar que el navegador no cachee nada tras el logout
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"]        = "no-cache"
    resp.headers["Expires"]       = "0"
    return resp


@app.route("/registro", methods=["GET", "POST"])
def registro():
    if request.method == "POST":
        nombre     = request.form.get("nombre", "").strip()
        correo     = request.form.get("correo", "").strip()
        contrasena = request.form.get("contrasena", "").strip()
        confirmar  = request.form.get("confirmar_contrasena", "").strip()

        if contrasena != confirmar:
            flash("Las contraseñas no coinciden.", "danger")
            return render_template("registro.html")
        if len(contrasena) < 6:
            flash("La contraseña debe tener al menos 6 caracteres.", "danger")
            return render_template("registro.html")
        # Verificar si el correo ya existe
        if db.query(Usuario).filter_by(correo=correo).first():
            flash("Ese correo ya está registrado.", "danger")
            return render_template("registro.html")

        try:
            hash_pw = bcrypt.hashpw(contrasena.encode(), bcrypt.gensalt()).decode()
            if correo.lower().endswith("@poligran.edu.co"):
                from database import Institucional
                nuevo = Institucional(nombre=nombre, correo=correo, contrasena=hash_pw)
            else:
                nuevo = PersonaExterna(nombre=nombre, correo=correo, contrasena=hash_pw)
            db.add(nuevo)
            db.commit()
            flash("Usuario registrado exitosamente. Inicia sesión.", "success")
            return redirect(url_for("login"))
        except Exception as e:
            db.rollback()
            flash(f"Error al registrar: {e}", "danger")

    return render_template("registro.html")


# ─── Rutas protegidas ──────────────────────────────────────────────────────

@app.route("/dashboard")
@requiere_login
def dashboard():
    u = usuario_logueado()
    # Datos para admin
    total_usuarios   = db.query(Usuario).count()
    total_pendientes = db.query(Reserva).filter_by(estado="pendiente").count() if u.rol == "admin" else 0
    total_aprobadas  = db.query(Reserva).filter_by(estado="aprobada").count()  if u.rol == "admin" else 0
    total_espacios   = db.query(Espacio).count()
    espacios         = db.query(Espacio).all()
    # Reservas del usuario actual (últimas 5)
    mis_reservas_rec = db.query(Reserva).filter_by(usuario_id=u.id)\
                         .order_by(Reserva.fecha.desc()).limit(5).all()
    # Conteo de mis reservas por estado
    mis_pendientes = db.query(Reserva).filter_by(usuario_id=u.id, estado="pendiente").count()
    mis_aprobadas  = db.query(Reserva).filter_by(usuario_id=u.id, estado="aprobada").count()
    mis_total      = db.query(Reserva).filter_by(usuario_id=u.id).count()
    return render_template("dashboard.html", usuario=u,
                           total_usuarios=total_usuarios,
                           total_pendientes=total_pendientes,
                           total_aprobadas=total_aprobadas,
                           total_espacios=total_espacios,
                           espacios=espacios,
                           mis_reservas_rec=mis_reservas_rec,
                           mis_pendientes=mis_pendientes,
                           mis_aprobadas=mis_aprobadas,
                           mis_total=mis_total)


@app.route("/perfil")
@requiere_login
def perfil():
    try:
        u = usuario_logueado()
        return render_template("perfil.html", usuario=u)
    except Exception:
        # Alt 3a: El sistema no logra conectar con la base de datos
        flash("Error al conectar con la base de datos. No se pudo cargar la información.", "danger")
        return redirect(url_for("dashboard"))


@app.route("/perfil/editar", methods=["GET", "POST"])
@requiere_login
def editar_perfil():
    u = usuario_logueado()

    if request.method == "POST":
        u.nombre = request.form.get("nombre", u.nombre).strip()

        # Cambio de contraseña opcional
        contrasena_actual  = request.form.get("contrasena_actual", "").strip()
        contrasena_nueva   = request.form.get("contrasena_nueva", "").strip()
        contrasena_confirmar = request.form.get("contrasena_confirmar", "").strip()

        if contrasena_nueva:
            # Verificar contraseña actual
            try:
                pw_ok = bcrypt.checkpw(contrasena_actual.encode(), u.contrasena.encode())
            except Exception:
                pw_ok = contrasena_actual == u.contrasena
            if not pw_ok:
                flash("La contraseña actual es incorrecta.", "danger")
                return render_template("editar_perfil.html", usuario=u)
            if contrasena_nueva != contrasena_confirmar:
                flash("La nueva contraseña y su confirmación no coinciden.", "danger")
                return render_template("editar_perfil.html", usuario=u)
            if len(contrasena_nueva) < 6:
                flash("La nueva contraseña debe tener al menos 6 caracteres.", "danger")
                return render_template("editar_perfil.html", usuario=u)
            u.contrasena = bcrypt.hashpw(contrasena_nueva.encode(), bcrypt.gensalt()).decode()

        if u.rol == "docente":
            u.cod_docente  = request.form.get("cod_docente", u.cod_docente).strip()
            u.departamento = request.form.get("departamento", u.departamento).strip()
        elif u.rol == "administrativo":
            u.dependencia = request.form.get("dependencia", u.dependencia).strip()
        elif u.rol == "externa":
            u.empresa  = request.form.get("empresa", u.empresa).strip()
            u.telefono = request.form.get("telefono", u.telefono).strip()

        try:
            db.commit()
            flash("Perfil actualizado correctamente.", "success")
            return redirect(url_for("perfil"))
        except Exception:
            # Alt 5a: El sistema no logra conectar con la base de datos
            db.rollback()
            flash("Error al conectar con la base de datos. No se guardaron los cambios.", "danger")
            return render_template("editar_perfil.html", usuario=u)

    return render_template("editar_perfil.html", usuario=u)


# ─── Rutas solo Admin ──────────────────────────────────────────────────────

@app.route("/admin/usuarios")
@requiere_login
@requiere_admin
def admin_usuarios():
    u      = usuario_logueado()
    rol    = request.args.get("rol", "")
    nombre = request.args.get("nombre", "").strip()
    query  = db.query(Usuario)
    if rol:
        query = query.filter_by(rol=rol)
    if nombre:
        query = query.filter(Usuario.nombre.ilike(f"%{nombre}%"))
    usuarios = query.all()
    return render_template("admin_usuarios.html", usuario=u, usuarios=usuarios,
                           filtro_rol=rol, filtro_nombre=nombre)


@app.route("/admin/usuarios/<int:uid>/editar-rol", methods=["GET", "POST"])
@requiere_login
@requiere_admin
def admin_editar_rol(uid):
    u        = usuario_logueado()
    objetivo = db.get(Usuario, uid)

    if not objetivo:
        flash("Usuario no encontrado.", "danger")
        return redirect(url_for("admin_usuarios"))

    if objetivo.id == u.id:
        flash("No puedes cambiarle el rol a tu propia cuenta.", "warning")
        return redirect(url_for("admin_usuarios"))

    if request.method == "POST":
        nuevo_rol = request.form.get("rol", "").strip()
        roles_validos = ["docente", "administrativo", "externa", "admin"]
        if nuevo_rol not in roles_validos:
            flash("Rol inválido.", "danger")
        else:
            objetivo.rol = nuevo_rol
            if nuevo_rol != "docente":
                objetivo.cod_docente  = None
                objetivo.departamento = None
            if nuevo_rol != "administrativo":
                objetivo.dependencia = None
            if nuevo_rol != "externa":
                objetivo.empresa  = None
                objetivo.telefono = None
            nombre_objetivo = objetivo.nombre
            db.commit()
            db.expire_all()
            flash(f"Rol de '{nombre_objetivo}' actualizado a '{nuevo_rol}' correctamente.", "success")
            return redirect(url_for("admin_usuarios"))

    return render_template("admin_editar_rol.html", usuario=u, objetivo=objetivo)


@app.route("/admin/usuarios/<int:uid>/desactivar", methods=["POST"])
@requiere_login
@requiere_admin
def admin_desactivar_usuario(uid):
    u        = usuario_logueado()
    objetivo = db.get(Usuario, uid)

    if not objetivo:
        flash("Usuario no encontrado.", "danger")
        return redirect(url_for("admin_usuarios"))

    if objetivo.id == u.id:
        flash("No puedes desactivar tu propia cuenta.", "warning")
        return redirect(url_for("admin_usuarios"))

    nombre = objetivo.nombre
    try:
        objetivo.activo = False
        db.commit()
        flash(f"Usuario '{nombre}' desactivado. Sus datos e historial se conservan.", "success")
    except Exception as e:
        db.rollback()
        flash(f"Error al desactivar el usuario: {e}", "danger")

    return redirect(url_for("admin_usuarios"))


@app.route("/admin/usuarios/<int:uid>/reactivar", methods=["POST"])
@requiere_login
@requiere_admin
def admin_reactivar_usuario(uid):
    u        = usuario_logueado()
    objetivo = db.get(Usuario, uid)

    if not objetivo:
        flash("Usuario no encontrado.", "danger")
        return redirect(url_for("admin_usuarios"))

    nombre = objetivo.nombre
    try:
        objetivo.activo = True
        db.commit()
        flash(f"Usuario '{nombre}' reactivado correctamente.", "success")
    except Exception as e:
        db.rollback()
        flash(f"Error al reactivar el usuario: {e}", "danger")

    return redirect(url_for("admin_usuarios"))


@app.route("/admin/usuarios/<int:uid>/strike", methods=["POST"])
@requiere_login
@requiere_admin
def admin_asignar_strike(uid):
    u        = usuario_logueado()
    objetivo = db.get(Usuario, uid)

    if not objetivo:
        flash("Usuario no encontrado.", "danger")
        return redirect(url_for("admin_usuarios"))

    if objetivo.id == u.id:
        flash("No puedes asignarte un strike a ti mismo.", "warning")
        return redirect(url_for("admin_usuarios"))

    try:
        objetivo.strikes = (objetivo.strikes or 0) + 1
        db.commit()
        flash(f"Strike asignado a '{objetivo.nombre}'. Total: {objetivo.strikes} strike(s).", "warning")
    except Exception as e:
        db.rollback()
        flash(f"Error al asignar el strike: {e}", "danger")

    return redirect(url_for("admin_usuarios"))


# ─── Punto de entrada (al final, después de registrar todas las rutas) ───────

# ══════════════════════════════════════════════════════════════════════════
#  SPRINT 5 — Agenda, Espacios y Reservas
# ══════════════════════════════════════════════════════════════════════════

from database import AgendaSemestral, BloqueHorario, Espacio, Reserva, Calificacion, Sancion
from datetime import datetime, time as dtime

# ── Restricciones horarias por defecto (RF-08, RF-09) ─────────────────────
# Horario permitido: 8:00 a.m. - 9:00 p.m.
HORA_APERTURA = dtime(8, 0)
HORA_CIERRE   = dtime(21, 0)

# Restricciones por día de semana:
# Lunes=0, Martes=1, Miércoles=2, Jueves=3, Viernes=4, Sábado=5, Domingo=6
# Lunes, miércoles y sábado → bloqueados TODO el día
DIAS_BLOQUEADOS_COMPLETO = [0, 2, 5]
# Martes y jueves → bloqueada la MAÑANA (hasta las 12:00)
DIAS_MANANA_BLOQUEADA    = [1, 3]
HORA_MANANA_FIN          = dtime(12, 0)
# Viernes → libre todo el día
# Domingo → no se puede reservar nunca

def horario_valido(fecha, hora_inicio, hora_fin):
    """Verifica que el horario cumpla las restricciones por defecto (RF-08, RF-09)."""
    errores = []
    dia = fecha.weekday()

    # Domingo: nunca permitido
    if dia == 6:
        errores.append("No se pueden hacer reservas los domingos.")
        return errores

    # Lunes, miércoles y sábado: bloqueados todo el día
    nombres_dias = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"]
    if dia in DIAS_BLOQUEADOS_COMPLETO:
        errores.append(f"Los {nombres_dias[dia]} no están disponibles para reservas.")
        return errores

    # Martes y jueves: mañana bloqueada (antes de las 12:00)
    if dia in DIAS_MANANA_BLOQUEADA:
        if hora_inicio < HORA_MANANA_FIN:
            errores.append(f"Los {nombres_dias[dia]} solo están disponibles desde las 12:00 p.m.")
            return errores

    # Horario general: 8:00 a.m. - 9:00 p.m.
    if hora_inicio < HORA_APERTURA:
        errores.append("La hora de inicio no puede ser antes de las 8:00 a.m.")
    if hora_fin > HORA_CIERRE:
        errores.append("La hora de fin no puede ser después de las 9:00 p.m.")
    if hora_inicio >= hora_fin:
        errores.append("La hora de inicio debe ser antes que la hora de fin.")

    return errores

def dia_bloqueado(fecha, hora_inicio, hora_fin):
    """Verifica si el horario cae en un bloque bloqueado."""
    agenda = db.query(AgendaSemestral).filter(
        AgendaSemestral.fecha_inicio <= fecha,
        AgendaSemestral.fecha_fin   >= fecha
    ).first()
    if not agenda:
        return False
    for bloque in agenda.bloques:
        if bloque.fecha == fecha:
            if not (hora_fin <= bloque.hora_inicio or hora_inicio >= bloque.hora_fin):
                return True
    return False

# ── Aplicar restricciones horarias por defecto (caso de uso 8) ────────────

@app.route("/admin/agenda/<int:aid>/aplicar-restricciones", methods=["POST"])
@requiere_login
@requiere_admin
def admin_aplicar_restricciones(aid):
    """
    Caso de uso 8: El administrador selecciona aplicar restricciones horarias.
    Pasos del diagrama:
    1. Admin selecciona aplicar restricciones
    2. El sistema aplica las reglas definidas
    3. El sistema realiza conexión con la base de datos
    4. El sistema guarda las restricciones aplicadas
    5. El sistema muestra confirmación
    """
    try:
        # Paso 3: El sistema realiza conexión con la base de datos
        agenda = db.get(AgendaSemestral, aid)
        if not agenda:
            flash("Agenda no encontrada.", "danger")
            return redirect(url_for("admin_agenda"))

        # Paso 2 y 4: Aplicar y guardar las restricciones en la BD
        restricciones_texto = (
            "Lunes/Miércoles/Sábado: bloqueados todo el día | "
            "Martes/Jueves: disponibles desde 12:00 p.m. | "
            "Viernes: 8:00 a.m. – 9:00 p.m."
        )
        agenda.restricciones_aplicadas = restricciones_texto
        db.commit()

        # Paso 5: El sistema muestra confirmación
        flash(
            f"Restricciones horarias aplicadas y guardadas correctamente: {restricciones_texto}",
            "success"
        )
    except Exception:
        # Alt 4a: El sistema no logra conectar con la base de datos
        db.rollback()
        flash("Error al conectar con la base de datos. No se guardaron las restricciones.", "danger")

    return redirect(url_for("admin_agenda"))


# ── AGENDA SEMESTRAL ──────────────────────────────────────────────────────

@app.route("/admin/agenda")
@requiere_login
@requiere_admin
def admin_agenda():
    u       = usuario_logueado()
    agendas = db.query(AgendaSemestral).order_by(AgendaSemestral.fecha_inicio.desc()).all()
    return render_template("admin_agenda.html", usuario=u, agendas=agendas)


@app.route("/admin/agenda/nueva", methods=["GET", "POST"])
@requiere_login
@requiere_admin
def admin_agenda_nueva():
    u = usuario_logueado()
    if request.method == "POST":
        semestre     = request.form.get("semestre", "").strip()
        fecha_inicio = request.form.get("fecha_inicio", "").strip()
        fecha_fin    = request.form.get("fecha_fin", "").strip()
        try:
            fi = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
            ff = datetime.strptime(fecha_fin,    "%Y-%m-%d").date()
            if fi >= ff:
                flash("La fecha de inicio debe ser antes que la fecha de fin.", "danger")
            else:
                try:
                    agenda = AgendaSemestral(semestre=semestre, fecha_inicio=fi, fecha_fin=ff)
                    db.add(agenda)
                    db.commit()
                    flash(f"Agenda '{semestre}' creada correctamente.", "success")
                    return redirect(url_for("admin_agenda"))
                except Exception:
                    # Alt 6a: El sistema no logra conectar con la base de datos
                    db.rollback()
                    flash("Error al conectar con la base de datos. No se guardó la agenda.", "danger")
        except ValueError:
            flash("Formato de fecha inválido.", "danger")
    return render_template("admin_agenda_nueva.html", usuario=u)


@app.route("/admin/agenda/<int:aid>/bloquear", methods=["GET", "POST"])
@requiere_login
@requiere_admin
def admin_bloquear_dia(aid):
    u      = usuario_logueado()
    agenda = db.get(AgendaSemestral, aid)
    if not agenda:
        flash("Agenda no encontrada.", "danger")
        return redirect(url_for("admin_agenda"))

    if request.method == "POST":
        fecha_str = request.form.get("fecha", "").strip()
        hi_str    = request.form.get("hora_inicio", "").strip()
        hf_str    = request.form.get("hora_fin", "").strip()
        try:
            fecha       = datetime.strptime(fecha_str, "%Y-%m-%d").date()
            hora_inicio = datetime.strptime(hi_str, "%H:%M").time()
            hora_fin    = datetime.strptime(hf_str, "%H:%M").time()

            if not (agenda.fecha_inicio <= fecha <= agenda.fecha_fin):
                flash("La fecha debe estar dentro del rango de la agenda.", "danger")
            elif hora_inicio >= hora_fin:
                flash("La hora de inicio debe ser antes que la hora de fin.", "danger")
            else:
                try:
                    bloque = BloqueHorario(agenda_id=agenda.id, fecha=fecha,
                                           hora_inicio=hora_inicio, hora_fin=hora_fin)
                    db.add(bloque)
                    db.commit()
                    flash("Bloque horario registrado correctamente.", "success")
                    return redirect(url_for("admin_agenda"))
                except Exception:
                    # Alt 6a: El sistema no logra conectar con la base de datos
                    db.rollback()
                    flash("Error al conectar con la base de datos. No se guardó el bloqueo.", "danger")
        except ValueError:
            flash("Formato de fecha u hora inválido.", "danger")

    return render_template("admin_bloquear_dia.html", usuario=u, agenda=agenda)


@app.route("/admin/agenda/<int:aid>/bloquear/<int:bid>/eliminar", methods=["POST"])
@requiere_login
@requiere_admin
def admin_eliminar_bloque(aid, bid):
    bloque = db.get(BloqueHorario, bid)
    if bloque and bloque.agenda_id == aid:
        db.delete(bloque)
        db.commit()
        flash("Bloque eliminado.", "success")
    return redirect(url_for("admin_agenda"))


# ── ESPACIOS ──────────────────────────────────────────────────────────────

@app.route("/admin/espacios")
@requiere_login
@requiere_admin
def admin_espacios():
    u        = usuario_logueado()
    espacios = db.query(Espacio).all()
    return render_template("admin_espacios.html", usuario=u, espacios=espacios)


@app.route("/admin/espacios/nuevo", methods=["GET", "POST"])
@requiere_login
@requiere_admin
def admin_espacio_nuevo():
    u = usuario_logueado()
    if request.method == "POST":
        nombre    = request.form.get("nombre", "").strip()
        capacidad = request.form.get("capacidad", "").strip()
        try:
            espacio = Espacio(nombre=nombre, capacidad=int(capacidad))
            db.add(espacio)
            db.commit()
            flash(f"Espacio '{nombre}' registrado correctamente.", "success")
            return redirect(url_for("admin_espacios"))
        except Exception as e:
            db.rollback()
            flash(f"Error al registrar: {e}", "danger")
    return render_template("admin_espacio_nuevo.html", usuario=u)


@app.route("/admin/espacios/<int:eid>/eliminar", methods=["POST"])
@requiere_login
@requiere_admin
def admin_espacio_eliminar(eid):
    espacio = db.get(Espacio, eid)
    if espacio:
        nombre = espacio.nombre
        db.delete(espacio)
        db.commit()
        flash(f"Espacio '{nombre}' eliminado.", "success")
    return redirect(url_for("admin_espacios"))


# ── RESERVAS ──────────────────────────────────────────────────────────────

@app.route("/reservas")
@requiere_login
def mis_reservas():
    u        = usuario_logueado()
    reservas = db.query(Reserva).filter_by(usuario_id=u.id)\
                 .order_by(Reserva.fecha.desc()).all()
    ahora = datetime.now()
    # IDs de reservas que ya finalizaron (para ocultar acciones en el template)
    finalizadas_ids = {r.id for r in reservas if reserva_ha_finalizado(r)}
    return render_template("mis_reservas.html", usuario=u, reservas=reservas,
                           ahora=ahora, finalizadas_ids=finalizadas_ids)


@app.route("/reservas/nueva", methods=["GET", "POST"])
@requiere_login
def reserva_nueva():
    u        = usuario_logueado()
    espacios = db.query(Espacio).all()

    if not espacios:
        flash("No hay espacios registrados. Contacta al administrador.", "warning")
        return redirect(url_for("mis_reservas"))

    # Verificar si el usuario tiene reservas pasadas sin calificar (cualquier tipo)
    reservas_usuario = db.query(Reserva).filter_by(usuario_id=u.id, estado="aprobada").all()
    pendientes_calificar = [
        r for r in reservas_usuario
        if reserva_ha_finalizado(r) and r.calificacion is None
    ]
    if pendientes_calificar:
        flash(
            f"Tienes {len(pendientes_calificar)} reserva(s) finalizada(s) sin calificar. "
            "Debes calificarla(s) antes de crear una nueva reserva.",
            "warning"
        )
        return redirect(url_for("historial_reservas"))

    if request.method == "POST":
        tipo_evento            = request.form.get("tipo_evento", "").strip()
        finalidad              = request.form.get("finalidad", "").strip()
        indumentaria           = request.form.get("indumentaria", "").strip()
        servicios_adicionales  = request.form.getlist("servicios_adicionales")
        cant_personas          = request.form.get("cant_personas", "").strip()
        espacio_id             = request.form.get("espacio_id", "").strip()
        fecha_str              = request.form.get("fecha", "").strip()
        hi_str                 = request.form.get("hora_inicio", "").strip()
        hf_str                 = request.form.get("hora_fin", "").strip()
        tiene_externos         = request.form.get("tiene_externos", "no") == "si"

        try:
            fecha       = datetime.strptime(fecha_str, "%Y-%m-%d").date()
            hora_inicio = datetime.strptime(hi_str, "%H:%M").time()
            hora_fin    = datetime.strptime(hf_str, "%H:%M").time()
            espacio     = db.get(Espacio, int(espacio_id))

            if not espacio:
                flash("Espacio no válido.", "danger")
            else:
                # Validar máx. 4 horas
                duracion = datetime.combine(fecha, hora_fin) - datetime.combine(fecha, hora_inicio)
                if duracion.total_seconds() > 4 * 3600:
                    flash("Una reserva no puede exceder 4 horas de duración.", "danger")
                elif horario_valido(fecha, hora_inicio, hora_fin):
                    for e in horario_valido(fecha, hora_inicio, hora_fin):
                        flash(e, "danger")
                elif dia_bloqueado(fecha, hora_inicio, hora_fin):
                    flash("Ese horario está bloqueado en la agenda semestral.", "danger")
                elif not espacio.consultar_disponibilidad(db, fecha, hora_inicio, hora_fin):
                    flash("El espacio no está disponible en ese horario.", "danger")
                elif int(cant_personas) > espacio.capacidad:
                    flash(f"El espacio solo tiene capacidad para {espacio.capacidad} personas.", "danger")
                else:
                    reserva = Reserva(
                        usuario_id            = u.id,
                        cant_personas         = int(cant_personas),
                        tipo_evento           = tipo_evento,
                        finalidad             = finalidad,
                        indumentaria          = indumentaria or None,
                        servicios_adicionales = ", ".join(servicios_adicionales) if servicios_adicionales else None,
                        fecha                 = fecha,
                        hora_inicio           = hora_inicio,
                        hora_fin              = hora_fin,
                        estado                = "pendiente",
                    )
                    reserva.espacios.append(espacio)
                    db.add(reserva)
                    db.commit()

                    # Registrar personas externas si aplica (método manual)
                    if tiene_externos:
                        nombres   = request.form.getlist("ext_nombre")
                        cedulas   = request.form.getlist("ext_cedula")
                        ingresos  = request.form.getlist("ext_fecha_ingreso")
                        salidas   = request.form.getlist("ext_fecha_salida")
                        for nombre, cedula, ing, sal in zip(nombres, cedulas, ingresos, salidas):
                            nombre  = nombre.strip()
                            cedula  = cedula.strip()
                            if not nombre or not cedula:
                                continue
                            fi = datetime.strptime(ing, "%Y-%m-%d").date() if ing.strip() else None
                            fs = datetime.strptime(sal, "%Y-%m-%d").date() if sal.strip() else None
                            db.add(PersonaExternaReserva(
                                reserva_id    = reserva.id,
                                nombre        = nombre,
                                cedula        = cedula,
                                fecha_ingreso = fi,
                                fecha_salida  = fs,
                            ))
                        db.commit()

                    enviar_correo_reserva_creada(u, reserva, espacio)
                    flash("Reserva solicitada correctamente. Queda pendiente de aprobación.", "success")
                    return redirect(url_for("mis_reservas"))
        except (ValueError, TypeError) as ex:
            flash(f"Datos inválidos. Revisa el formulario.", "danger")
        except Exception as ex:
            db.rollback()
            traceback.print_exc()
            flash(f"Error al procesar la reserva: {ex}", "danger")
        except Exception as ex:
            db.rollback()
            raise ex
    return render_template("reserva_nueva.html", usuario=u, espacios=espacios)



# ── Recuperación de contraseña ──────────────────────────────────────────────

@app.route("/recuperar-contrasena", methods=["GET", "POST"])
def recuperar_contrasena():
    if request.method == "POST":
        correo = request.form.get("correo", "").strip()
        usuario = db.query(Usuario).filter_by(correo=correo).first()
        if usuario:
            token = serializer.dumps(correo, salt="recuperar-contrasena")
            enviado = enviar_correo_recuperacion(correo, usuario.nombre, token)
            if enviado:
                flash("Te enviamos un enlace de recuperación a tu correo.", "success")
            else:
                flash("Error al enviar el correo. Intenta de nuevo.", "danger")
        else:
            flash("Si el correo existe, recibirás las instrucciones.", "info")
        return redirect(url_for("recuperar_contrasena"))
    return render_template("recuperar_contrasena.html")


@app.route("/recuperar-contrasena/<token>", methods=["GET", "POST"])
def recuperar_contrasena_reset(token):
    try:
        correo = serializer.loads(token, salt="recuperar-contrasena", max_age=1800)
    except Exception:
        flash("El enlace de recuperación es inválido o ha expirado.", "danger")
        return redirect(url_for("login"))
    usuario = db.query(Usuario).filter_by(correo=correo).first()
    if not usuario:
        flash("Usuario no encontrado.", "danger")
        return redirect(url_for("login"))
    if request.method == "POST":
        nueva = request.form.get("contrasena_nueva", "").strip()
        confirmar = request.form.get("contrasena_confirmar", "").strip()
        if len(nueva) < 6:
            flash("La contraseña debe tener al menos 6 caracteres.", "danger")
        elif nueva != confirmar:
            flash("Las contraseñas no coinciden.", "danger")
        else:
            usuario.contrasena = bcrypt.hashpw(nueva.encode(), bcrypt.gensalt()).decode()
            db.commit()
            flash("Contraseña actualizada correctamente. Ya puedes iniciar sesión.", "success")
            return redirect(url_for("login"))
    return render_template("recuperar_contrasena_reset.html", token=token)


# ── Admin: gestión de reservas ──────────────────────────────────────────────

@app.route("/reservas/<int:rid>/ver")
@requiere_login
def ver_reserva(rid):
    u = usuario_logueado()
    reserva = db.get(Reserva, rid)

    if not reserva:
        flash("Reserva no encontrada.", "danger")
        return redirect(url_for("admin_historial") if u.rol == "admin" else url_for("mis_reservas"))

    if u.rol != "admin" and reserva.usuario_id != u.id:
        flash("No tienes permiso para ver esta reserva.", "danger")
        return redirect(url_for("mis_reservas"))

    return render_template("ver_reserva.html", usuario=u, reserva=reserva)


@app.route("/admin/sanciones")
@requiere_login
@requiere_admin
def admin_sanciones():
    u = usuario_logueado()
    sanciones = (
        db.query(Sancion)
        .order_by(Sancion.fecha.desc(), Sancion.id.desc())
        .all()
    )
    return render_template("admin_sanciones.html", usuario=u, sanciones=sanciones)


@app.route("/admin/sanciones/nueva", methods=["GET", "POST"])
@requiere_login
@requiere_admin
def admin_sancion_nueva():
    u = usuario_logueado()
    usuarios = db.query(Usuario).order_by(Usuario.nombre.asc()).all()
    reservas = db.query(Reserva).order_by(Reserva.fecha.desc()).all()

    if request.method == "POST":
        usuario_id = request.form.get("usuario_id", type=int)
        reserva_id = request.form.get("reserva_id", type=int)
        tipo = request.form.get("tipo", "advertencia").strip()
        descripcion = request.form.get("descripcion", "").strip()

        if tipo not in ["advertencia", "suspension"]:
            flash("Tipo de sanción inválido.", "danger")
            return render_template("admin_sancion_nueva.html", usuario=u, usuarios=usuarios, reservas=reservas)
        if not usuario_id or not db.get(Usuario, usuario_id):
            flash("Debes seleccionar un usuario válido.", "danger")
            return render_template("admin_sancion_nueva.html", usuario=u, usuarios=usuarios, reservas=reservas)
        if reserva_id and not db.get(Reserva, reserva_id):
            flash("La reserva asociada no existe.", "danger")
            return render_template("admin_sancion_nueva.html", usuario=u, usuarios=usuarios, reservas=reservas)
        if not descripcion:
            flash("La descripción de la sanción es obligatoria.", "danger")
            return render_template("admin_sancion_nueva.html", usuario=u, usuarios=usuarios, reservas=reservas)

        try:
            sancion = Sancion(
                usuario_id=usuario_id,
                reserva_id=reserva_id,
                admin_id=u.id,
                tipo=tipo,
                descripcion=descripcion,
                fecha=date.today(),
            )
            db.add(sancion)
            db.commit()
            flash("Sanción registrada correctamente.", "success")
            return redirect(url_for("admin_sanciones"))
        except Exception as e:
            db.rollback()
            flash(f"Error al registrar la sanción: {e}", "danger")

    return render_template("admin_sancion_nueva.html", usuario=u, usuarios=usuarios, reservas=reservas)


@app.route("/admin/reservas")
@requiere_login
@requiere_admin
def admin_reservas():
    from database import Reserva
    u = usuario_logueado()
    reservas = db.query(Reserva).filter_by(estado="pendiente").all()
    return render_template("admin_reservas.html", usuario=u, reservas=reservas)


@app.route("/admin/reservas/<int:rid>/aprobar", methods=["GET", "POST"])
@requiere_login
@requiere_admin
def admin_aprobar_reserva(rid):
    from database import Reserva
    u       = usuario_logueado()
    reserva = db.get(Reserva, rid)

    if not reserva:
        flash("Reserva no encontrada.", "danger")
        return redirect(url_for("admin_reservas"))

    if request.method == "POST":
        try:
            # Paso 4 del diagrama: admin confirma la aprobación
            # Paso 5: El sistema realiza conexión con la base de datos
            reserva.estado = "aprobada"
            db.commit()
            usuario_reserva = db.get(Usuario, reserva.usuario_id)
            espacio = reserva.espacios[0] if reserva.espacios else None
            if usuario_reserva and espacio:
                enviar_correo_reserva_aprobada(usuario_reserva, reserva, espacio)
            flash("Reserva confirmada y notificación enviada al usuario.", "success")
        except Exception:
            # Alt 5a: El sistema no logra conectar con la base de datos
            db.rollback()
            flash("Error al conectar con la base de datos. No se actualizó el estado.", "danger")
        return redirect(url_for("admin_reservas"))

    # GET: mostrar detalles antes de confirmar (paso 3 del diagrama)
    return render_template("admin_aprobar_reserva.html", usuario=u, reserva=reserva)


@app.route("/admin/reservas/<int:rid>/rechazar", methods=["GET", "POST"])
@requiere_login
@requiere_admin
def admin_rechazar_reserva(rid):
    from database import Reserva
    u       = usuario_logueado()
    reserva = db.get(Reserva, rid)

    if not reserva:
        flash("Reserva no encontrada.", "danger")
        return redirect(url_for("admin_reservas"))

    if request.method == "POST":
        justificacion = request.form.get("justificacion", "").strip()

        # Paso 4 del diagrama: el admin DEBE ingresar justificación
        if not justificacion:
            flash("Debes ingresar una justificación para rechazar la reserva.", "danger")
            return render_template("admin_rechazar_reserva.html", usuario=u, reserva=reserva)

        try:
            # Paso 7-8: actualizar estado y guardar justificación
            # Paso 6: El sistema realiza conexión con la base de datos
            reserva.estado = "rechazada"
            reserva.justificacion_rechazo = justificacion
            db.commit()
            # Notificar al usuario
            usuario_reserva = db.get(Usuario, reserva.usuario_id)
            espacio = reserva.espacios[0] if reserva.espacios else None
            if usuario_reserva and espacio:
                enviar_correo_reserva_rechazada(usuario_reserva, reserva, espacio, justificacion)
            flash("Reserva rechazada. Se notificó al usuario.", "info")
        except Exception:
            # Alt 6a: El sistema no logra conectar con la base de datos
            db.rollback()
            flash("Error al conectar con la base de datos. No se guardaron los cambios.", "danger")
        return redirect(url_for("admin_reservas"))

    # GET: mostrar formulario con detalles de la reserva
    return render_template("admin_rechazar_reserva.html", usuario=u, reserva=reserva)


# ── Sprint 6: Cancelar y modificar reserva ─────────────────────────────────

@app.route("/reservas/<int:rid>/cancelar", methods=["POST"])
@requiere_login
def cancelar_reserva(rid):
    from database import Reserva
    from datetime import datetime
    u = usuario_logueado()
    reserva = db.get(Reserva, rid)
    if not reserva or reserva.usuario_id != u.id:
        flash("Reserva no encontrada.", "danger")
        return redirect(url_for("mis_reservas"))
    # Bloquear si la reserva ya finalizó
    if reserva_ha_finalizado(reserva):
        flash("No puedes cancelar una reserva que ya finalizó.", "danger")
        return redirect(url_for("mis_reservas"))
    if reserva.estado not in ["pendiente", "aprobada"]:
        flash("Esta reserva no puede cancelarse.", "warning")
        return redirect(url_for("mis_reservas"))
    # RF-12: mínimo 4 horas de anticipación
    from datetime import datetime, time as dtime2
    ahora = datetime.now()
    fecha_hora_reserva = datetime.combine(reserva.fecha, reserva.hora_inicio)
    if (fecha_hora_reserva - ahora).total_seconds() < 4 * 3600:
        flash("No puedes cancelar una reserva con menos de 4 horas de anticipación.", "danger")
        return redirect(url_for("mis_reservas"))
    try:
        espacio = reserva.espacios[0] if reserva.espacios else None
        reserva.estado = "cancelada"
        db.commit()
        if espacio:
            enviar_correo_reserva_cancelada(u, reserva, espacio)
        flash("Reserva cancelada correctamente. Se envió una confirmación a tu correo.", "success")
    except Exception:
        db.rollback()
        flash("Error al conectar con la base de datos. No se realizó la cancelación.", "danger")
    return redirect(url_for("mis_reservas"))


@app.route("/reservas/<int:rid>/modificar", methods=["GET", "POST"])
@requiere_login
def modificar_reserva(rid):
    from database import Reserva
    from datetime import datetime, date as ddate, time as dtime2
    u = usuario_logueado()
    reserva = db.get(Reserva, rid)
    if not reserva or reserva.usuario_id != u.id:
        flash("Reserva no encontrada.", "danger")
        return redirect(url_for("mis_reservas"))
    # Bloquear si la reserva ya finalizó
    if reserva_ha_finalizado(reserva):
        flash("No puedes modificar una reserva que ya finalizó.", "danger")
        return redirect(url_for("mis_reservas"))
    if reserva.estado != "pendiente":
        flash("Solo puedes modificar reservas que estén En revisión.", "warning")
        return redirect(url_for("mis_reservas"))
    # RF-13: mínimo 1 semana de anticipación
    ahora = datetime.now()
    fecha_hora_reserva = datetime.combine(reserva.fecha, reserva.hora_inicio)
    if (fecha_hora_reserva - ahora).total_seconds() < 7 * 24 * 3600:
        flash("Solo puedes modificar una reserva con al menos 1 semana de anticipación.", "danger")
        return redirect(url_for("mis_reservas"))
    if request.method == "POST":
        try:
            nueva_fecha    = ddate.fromisoformat(request.form.get("fecha"))
            nueva_hora_ini = dtime2.fromisoformat(request.form.get("hora_inicio"))
            nueva_hora_fin = dtime2.fromisoformat(request.form.get("hora_fin"))
            errores = horario_valido(nueva_fecha, nueva_hora_ini, nueva_hora_fin)
            if errores:
                for e in errores:
                    flash(e, "danger")
                return render_template("modificar_reserva.html", reserva=reserva, usuario=u)
            reserva.fecha       = nueva_fecha
            reserva.hora_inicio = nueva_hora_ini
            reserva.hora_fin    = nueva_hora_fin
            try:
                db.commit()
                flash("Reserva modificada correctamente. Sigue en revisión.", "success")
                return redirect(url_for("mis_reservas"))
            except Exception:
                db.rollback()
                flash("Error al conectar con la base de datos. No se guardaron los cambios.", "danger")
        except Exception as e:
            flash(f"Error al modificar: {e}", "danger")
    return render_template("modificar_reserva.html", reserva=reserva, usuario=u)


# ── Admin: modificar cualquier reserva ─────────────────────────────────────

@app.route("/admin/reservas/<int:rid>/modificar", methods=["GET", "POST"])
@requiere_login
@requiere_admin
def admin_modificar_reserva(rid):
    from database import Reserva
    from datetime import date as ddate, time as dtime2
    u = usuario_logueado()
    reserva = db.get(Reserva, rid)
    if not reserva:
        flash("Reserva no encontrada.", "danger")
        return redirect(url_for("admin_reservas"))
    # Bloquear si la reserva ya finalizó
    if reserva_ha_finalizado(reserva):
        flash("No se puede modificar una reserva que ya finalizó.", "danger")
        return redirect(url_for("admin_historial"))
    if reserva.estado not in ["pendiente", "aprobada"]:
        flash("Solo se pueden modificar reservas en revisión o aprobadas.", "warning")
        return redirect(url_for("admin_reservas"))
    if request.method == "POST":
        try:
            nueva_fecha    = ddate.fromisoformat(request.form.get("fecha"))
            nueva_hora_ini = dtime2.fromisoformat(request.form.get("hora_inicio"))
            nueva_hora_fin = dtime2.fromisoformat(request.form.get("hora_fin"))
            errores = horario_valido(nueva_fecha, nueva_hora_ini, nueva_hora_fin)
            if errores:
                for e in errores:
                    flash(e, "danger")
                return render_template("admin_modificar_reserva.html", reserva=reserva, usuario=u)
            reserva.fecha       = nueva_fecha
            reserva.hora_inicio = nueva_hora_ini
            reserva.hora_fin    = nueva_hora_fin
            try:
                db.commit()
                flash("Reserva modificada correctamente.", "success")
                return redirect(url_for("admin_reservas", estado=reserva.estado))
            except Exception:
                db.rollback()
                flash("Error al conectar con la base de datos. No se guardaron los cambios.", "danger")
        except Exception as e:
            flash(f"Error al modificar: {e}", "danger")
    return render_template("admin_modificar_reserva.html", reserva=reserva, usuario=u)


# ── Calendario público de reservas ──────────────────────────────────────────

COLORES_ESPACIOS = {
    "Sala B1":               "#00b4d8",
    "Sala B2":               "#0077a8",
    "Sala B3":               "#4361ee",
    "Sala de Música y Baile": "#7b2d8b",
    "Auditorio Completo":    "#e06c00",
}

@app.route("/calendario")
@requiere_login
def calendario():
    u = usuario_logueado()
    espacios = db.query(Espacio).all()
    return render_template("calendario.html", usuario=u, espacios=espacios,
                           colores=COLORES_ESPACIOS)


@app.route("/api/eventos-calendario")
@requiere_login
def api_eventos_calendario():
    reservas = db.query(Reserva).filter(
        Reserva.estado.in_(["pendiente", "aprobada"])
    ).all()
    eventos = []
    for r in reservas:
        for esp in r.espacios:
            color = COLORES_ESPACIOS.get(esp.nombre, "#6c757d")
            titulo = f"{esp.nombre} — {r.tipo_evento}"
            if r.estado == "pendiente":
                titulo = "⏳ " + titulo
            else:
                titulo = "✅ " + titulo
            eventos.append({
                "id":              str(r.id),
                "title":           titulo,
                "start":           f"{r.fecha}T{r.hora_inicio.strftime('%H:%M:%S')}",
                "end":             f"{r.fecha}T{r.hora_fin.strftime('%H:%M:%S')}",
                "backgroundColor": color if r.estado == "aprobada" else color + "99",
                "borderColor":     color,
                "textColor":       "#ffffff",
                "extendedProps": {
                    "espacio":      esp.nombre,
                    "tipo_evento":  r.tipo_evento,
                    "finalidad":    r.finalidad or "",
                    "estado":       r.estado,
                    "personas":     r.cant_personas,
                    "hora_inicio":  r.hora_inicio.strftime("%H:%M"),
                    "hora_fin":     r.hora_fin.strftime("%H:%M"),
                }
            })
    return jsonify(eventos)


@app.route("/api/disponibilidad")
@requiere_login
def api_disponibilidad():
    espacio_id = request.args.get("espacio_id", type=int)
    fecha_str  = request.args.get("fecha", "")
    if not espacio_id or not fecha_str:
        return jsonify({"ocupados": []})
    try:
        fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"ocupados": []})
    espacio = db.get(Espacio, espacio_id)
    if not espacio:
        return jsonify({"ocupados": []})
    # Incluir conflictos físicos: si B1 está ocupado, bloquear también B3 y Auditorio Completo
    nombres_verificar = [espacio.nombre] + CONFLICTOS.get(espacio.nombre, [])
    vistos = set()
    ocupados = []
    for nombre in nombres_verificar:
        esp_c = db.query(Espacio).filter_by(nombre=nombre).first()
        if not esp_c:
            continue
        reservas_ocup = db.query(Reserva).filter(
            Reserva.espacios.contains(esp_c),
            Reserva.fecha == fecha,
            Reserva.estado.in_(["aprobada", "pendiente"])
        ).all()
        for r in reservas_ocup:
            key = (r.hora_inicio, r.hora_fin)
            if key not in vistos:
                vistos.add(key)
                ocupados.append({
                    "inicio": r.hora_inicio.strftime("%H:%M"),
                    "fin":    r.hora_fin.strftime("%H:%M")
                })
    return jsonify({"ocupados": ocupados})



# ── Crear reserva desde el panel admin (Sprint 7 - caso 13) ────────────────

@app.route("/admin/reservas/nueva", methods=["GET", "POST"])
@requiere_login
@requiere_admin
def admin_crear_reserva():
    from database import Reserva, Espacio
    u        = usuario_logueado()
    espacios = db.query(Espacio).all()
    usuarios_lista = db.query(Usuario).all()

    if request.method == "POST":
        try:
            usuario_id    = int(request.form.get("usuario_id"))
            espacio_id    = int(request.form.get("espacio_id"))
            cant_personas = int(request.form.get("cant_personas"))
            tipo_evento   = request.form.get("tipo_evento","").strip()
            finalidad     = request.form.get("finalidad","").strip()
            indumentaria  = request.form.get("indumentaria","").strip()
            servicios     = request.form.getlist("servicios_adicionales")
            fecha         = date.fromisoformat(request.form.get("fecha"))
            hora_inicio   = dtime.fromisoformat(request.form.get("hora_inicio"))
            hora_fin      = dtime.fromisoformat(request.form.get("hora_fin"))

            espacio = db.get(Espacio, espacio_id)
            if not espacio:
                flash("Espacio no encontrado.", "danger")
                return render_template("admin_crear_reserva.html", usuario=u,
                                       espacios=espacios, usuarios_lista=usuarios_lista)

            # Validar máx. 4 horas
            from datetime import datetime as _dt
            duracion = _dt.combine(fecha, hora_fin) - _dt.combine(fecha, hora_inicio)
            if duracion.total_seconds() > 4 * 3600:
                flash("Una reserva no puede exceder 4 horas de duración.", "danger")
                return render_template("admin_crear_reserva.html", usuario=u,
                                       espacios=espacios, usuarios_lista=usuarios_lista)
            try:
                reserva = Reserva(
                    usuario_id            = usuario_id,
                    cant_personas         = cant_personas,
                    tipo_evento           = tipo_evento,
                    finalidad             = finalidad or None,
                    indumentaria          = indumentaria or None,
                    servicios_adicionales = ", ".join(servicios) if servicios else None,
                    fecha                 = fecha,
                    hora_inicio           = hora_inicio,
                    hora_fin              = hora_fin,
                    estado                = "aprobada",
                )
                reserva.espacios.append(espacio)
                db.add(reserva)
                db.commit()
                flash("Reserva creada y confirmada correctamente.", "success")
                return redirect(url_for("admin_historial"))
            except Exception:
                db.rollback()
                flash("Error al conectar con la base de datos. No se registró la reserva.", "danger")
        except Exception as e:
            flash(f"Datos inválidos: {e}", "danger")

    return render_template("admin_crear_reserva.html", usuario=u,
                           espacios=espacios, usuarios_lista=usuarios_lista)


# ── Personas externas — Subida de Excel ────────────────────────────────────

@app.route("/reservas/<int:rid>/personas-externas/excel", methods=["POST"])
@requiere_login
def personas_externas_excel(rid):
    """Carga personas externas desde un archivo Excel (.xlsx / .xls)."""
    from database import Reserva
    u = usuario_logueado()
    reserva = db.get(Reserva, rid)

    if not reserva or reserva.usuario_id != u.id:
        flash("Reserva no encontrada.", "danger")
        return redirect(url_for("mis_reservas"))

    archivo = request.files.get("archivo_excel")
    if not archivo or archivo.filename == "":
        flash("No se seleccionó ningún archivo.", "danger")
        return redirect(url_for("mis_reservas"))

    nombre_archivo = archivo.filename.lower()
    if not (nombre_archivo.endswith(".xlsx") or nombre_archivo.endswith(".xls") or nombre_archivo.endswith(".csv")):
        flash("El archivo debe ser formato Excel (.xlsx o .xls) o CSV (.csv).", "danger")
        return redirect(url_for("mis_reservas"))

    def parsear_fecha(val):
        if val is None:
            return None
        if hasattr(val, 'date'):
            return val.date() if hasattr(val, 'hour') else val
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(str(val).strip(), fmt).date()
            except ValueError:
                continue
        return None

    try:
        agregados = 0
        errores   = []
        filas     = []

        if nombre_archivo.endswith(".csv"):
            import csv, io
            contenido = archivo.read().decode("utf-8-sig")  # utf-8-sig elimina BOM si existe
            reader = csv.reader(io.StringIO(contenido), delimiter=";")
            # intentar con coma si punto y coma no da 4 columnas
            primera_linea = contenido.split("\n")[0]
            if primera_linea.count(";") < 3 and primera_linea.count(",") >= 3:
                reader = csv.reader(io.StringIO(contenido), delimiter=",")
            rows = list(reader)
            expected = ["nombre", "cedula", "fecha_ingreso", "fecha_salida"]
            headers = [c.strip().lower() for c in rows[0]] if rows else []
            if headers[:4] != expected:
                flash(f"Estructura incorrecta. La primera fila debe contener: {', '.join(expected)}", "danger")
                return redirect(url_for("mis_reservas"))
            for fila_num, row in enumerate(rows[1:], start=2):
                if len(row) < 2:
                    continue
                filas.append((fila_num, row[0], row[1], row[2] if len(row) > 2 else None, row[3] if len(row) > 3 else None))
        else:
            import openpyxl
            from io import BytesIO
            wb = openpyxl.load_workbook(BytesIO(archivo.read()), data_only=True)
            ws = wb.active
            headers = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, 5)]
            expected = ["nombre", "cedula", "fecha_ingreso", "fecha_salida"]
            if headers[:4] != expected:
                flash(f"Estructura incorrecta. La primera fila debe contener: {', '.join(expected)}", "danger")
                return redirect(url_for("mis_reservas"))
            for fila_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                filas.append((fila_num, row[0], row[1], row[2], row[3]))

        for fila_num, nombre_p, cedula_p, ing_raw, sal_raw in filas:
            nombre_p = str(nombre_p or "").strip()
            cedula_p = str(cedula_p or "").strip()

            if not nombre_p and not cedula_p:
                continue
            if not nombre_p or not cedula_p:
                errores.append(f"Fila {fila_num}: nombre y cédula son obligatorios.")
                continue

            fi = parsear_fecha(ing_raw)
            fs = parsear_fecha(sal_raw)

            db.add(PersonaExternaReserva(
                reserva_id    = reserva.id,
                nombre        = nombre_p,
                cedula        = cedula_p,
                fecha_ingreso = fi,
                fecha_salida  = fs,
            ))
            agregados += 1

        db.commit()

        if errores:
            for err in errores:
                flash(err, "warning")
        if agregados:
            flash(f"{agregados} persona(s) externa(s) agregada(s) correctamente.", "success")
        else:
            flash("No se encontraron registros válidos en el archivo.", "warning")

    except Exception as e:
        db.rollback()
        flash(f"Error al procesar el archivo Excel: {e}", "danger")

    return redirect(url_for("mis_reservas"))


# ── Eliminar reserva desde el panel admin ──────────────────────────────────

@app.route("/admin/reservas/<int:rid>/eliminar", methods=["POST"])
@requiere_login
@requiere_admin
def admin_eliminar_reserva(rid):
    from database import Reserva
    reserva = db.get(Reserva, rid)
    if not reserva:
        flash("Reserva no encontrada.", "danger")
        return redirect(url_for("admin_historial"))
    # No se pueden eliminar reservas ya finalizadas
    if reserva_ha_finalizado(reserva):
        flash("No se pueden eliminar reservas que ya finalizaron.", "danger")
        return redirect(url_for("admin_historial"))
    try:
        db.delete(reserva)
        db.commit()
        flash("Reserva eliminada correctamente.", "success")
    except Exception:
        db.rollback()
        flash("Error al conectar con la base de datos. No se eliminó la reserva.", "danger")
    return redirect(url_for("admin_historial"))



# ══════════════════════════════════════════════════════════════════════════
#  HISTORIAL COMPLETO DE RESERVAS — Usuario
# ══════════════════════════════════════════════════════════════════════════

@app.route("/reservas/historial")
@requiere_login
def historial_reservas():
    """Historial completo con filtros por estado, fecha y texto libre."""
    u = usuario_logueado()

    estado     = request.args.get("estado", "")
    fecha_desde = request.args.get("fecha_desde", "")
    fecha_hasta = request.args.get("fecha_hasta", "")
    busqueda    = request.args.get("busqueda", "").strip()

    query = db.query(Reserva).filter_by(usuario_id=u.id)

    if estado:
        query = query.filter(Reserva.estado == estado)

    if fecha_desde:
        try:
            fd = datetime.strptime(fecha_desde, "%Y-%m-%d").date()
            query = query.filter(Reserva.fecha >= fd)
        except ValueError:
            pass

    if fecha_hasta:
        try:
            fh = datetime.strptime(fecha_hasta, "%Y-%m-%d").date()
            query = query.filter(Reserva.fecha <= fh)
        except ValueError:
            pass

    if busqueda:
        query = query.filter(
            Reserva.tipo_evento.ilike(f"%{busqueda}%") |
            Reserva.finalidad.ilike(f"%{busqueda}%")
        )

    reservas = query.order_by(Reserva.fecha.desc()).all()

    # Conteos para las pestañas
    conteos = {
        "todas":     db.query(Reserva).filter_by(usuario_id=u.id).count(),
        "pendiente": db.query(Reserva).filter_by(usuario_id=u.id, estado="pendiente").count(),
        "aprobada":  db.query(Reserva).filter_by(usuario_id=u.id, estado="aprobada").count(),
        "rechazada": db.query(Reserva).filter_by(usuario_id=u.id, estado="rechazada").count(),
        "cancelada": db.query(Reserva).filter_by(usuario_id=u.id, estado="cancelada").count(),
    }

    hoy = date.today()
    ahora = datetime.now()
    # IDs de reservas ya finalizadas (para el template)
    finalizadas_ids = {r.id for r in reservas if reserva_ha_finalizado(r)}
    return render_template(
        "historial_reservas.html",
        usuario=u,
        reservas=reservas,
        conteos=conteos,
        hoy=hoy,
        ahora=ahora,
        finalizadas_ids=finalizadas_ids,
        filtro_estado=estado,
        filtro_fecha_desde=fecha_desde,
        filtro_fecha_hasta=fecha_hasta,
        filtro_busqueda=busqueda,
    )


# ══════════════════════════════════════════════════════════════════════════
#  CALIFICACIONES — Usuario
# ══════════════════════════════════════════════════════════════════════════

@app.route("/reservas/<int:rid>/calificar", methods=["GET", "POST"])
@requiere_login
def calificar_reserva(rid):
    """El usuario califica una reserva aprobada cuya fecha ya pasó."""
    u = usuario_logueado()
    reserva = db.get(Reserva, rid)

    if not reserva or reserva.usuario_id != u.id:
        flash("Reserva no encontrada.", "danger")
        return redirect(url_for("historial_reservas"))

    if reserva.estado != "aprobada":
        flash("Solo puedes calificar reservas confirmadas.", "warning")
        return redirect(url_for("historial_reservas"))

    hoy = date.today()
    ahora = datetime.now()
    fecha_hora_fin = datetime.combine(reserva.fecha, reserva.hora_fin)
    if ahora <= fecha_hora_fin:
        flash("Solo puedes calificar reservas que ya han finalizado.", "warning")
        return redirect(url_for("historial_reservas"))

    # Verificar si ya tiene calificación
    cal_existente = reserva.calificacion

    if request.method == "POST":
        try:
            puntuacion = int(request.form.get("puntuacion", 0))
            comentario = request.form.get("comentario", "").strip()

            if puntuacion < 1 or puntuacion > 5:
                flash("Selecciona una puntuación entre 1 y 5 estrellas.", "danger")
                return render_template("calificar_reserva.html",
                                       usuario=u, reserva=reserva,
                                       calificacion=cal_existente)

            if cal_existente:
                # Editar calificación existente
                cal_existente.puntuacion = puntuacion
                cal_existente.comentario = comentario
                cal_existente.fecha = hoy
                db.commit()
                flash("Calificación actualizada correctamente.", "success")
            else:
                # Nueva calificación
                nueva_cal = Calificacion(
                    reserva_id=reserva.id,
                    puntuacion=puntuacion,
                    comentario=comentario,
                    fecha=hoy,
                )
                db.add(nueva_cal)
                db.commit()
                flash("¡Gracias por tu calificación!", "success")

            return redirect(url_for("historial_reservas"))

        except (ValueError, TypeError):
            flash("Datos inválidos. Intenta de nuevo.", "danger")

    return render_template("calificar_reserva.html",
                           usuario=u, reserva=reserva,
                           calificacion=cal_existente)


# ══════════════════════════════════════════════════════════════════════════
#  HISTORIAL COMPLETO DE RESERVAS — Admin
# ══════════════════════════════════════════════════════════════════════════

@app.route("/admin/historial")
@requiere_login
@requiere_admin
def admin_historial():
    """Panel admin: historial completo con filtros y estadísticas."""
    u = usuario_logueado()

    estado      = request.args.get("estado", "")
    fecha_desde = request.args.get("fecha_desde", "")
    fecha_hasta = request.args.get("fecha_hasta", "")
    busqueda    = request.args.get("busqueda", "").strip()
    espacio_id  = request.args.get("espacio_id", "")

    query = db.query(Reserva)

    if estado:
        query = query.filter(Reserva.estado == estado)

    if fecha_desde:
        try:
            fd = datetime.strptime(fecha_desde, "%Y-%m-%d").date()
            query = query.filter(Reserva.fecha >= fd)
        except ValueError:
            pass

    if fecha_hasta:
        try:
            fh = datetime.strptime(fecha_hasta, "%Y-%m-%d").date()
            query = query.filter(Reserva.fecha <= fh)
        except ValueError:
            pass

    if busqueda:
        query = query.join(Reserva.usuario).filter(
            Usuario.nombre.ilike(f"%{busqueda}%") |
            Usuario.correo.ilike(f"%{busqueda}%") |
            Reserva.tipo_evento.ilike(f"%{busqueda}%")
        )

    if espacio_id:
        try:
            esp = db.get(Espacio, int(espacio_id))
            if esp:
                query = query.filter(Reserva.espacios.contains(esp))
        except (ValueError, TypeError):
            pass

    reservas = query.order_by(Reserva.fecha.desc()).all()
    espacios = db.query(Espacio).all()

    # Estadísticas globales
    total_reservas  = db.query(Reserva).count()
    total_aprobadas = db.query(Reserva).filter_by(estado="aprobada").count()
    total_pend      = db.query(Reserva).filter_by(estado="pendiente").count()
    total_rechaz    = db.query(Reserva).filter_by(estado="rechazada").count()
    total_cancel    = db.query(Reserva).filter_by(estado="cancelada").count()

    # Promedio de calificaciones
    cals = db.query(Calificacion).all()
    promedio_global = round(sum(c.puntuacion for c in cals) / len(cals), 1) if cals else None

    # IDs de reservas ya finalizadas (para ocultar acciones en el template)
    finalizadas_ids = {r.id for r in reservas if reserva_ha_finalizado(r)}

    return render_template(
        "admin_historial.html",
        usuario=u,
        reservas=reservas,
        espacios=espacios,
        finalizadas_ids=finalizadas_ids,
        total_reservas=total_reservas,
        total_aprobadas=total_aprobadas,
        total_pend=total_pend,
        total_rechaz=total_rechaz,
        total_cancel=total_cancel,
        promedio_global=promedio_global,
        filtro_estado=estado,
        filtro_fecha_desde=fecha_desde,
        filtro_fecha_hasta=fecha_hasta,
        filtro_busqueda=busqueda,
        filtro_espacio_id=espacio_id,
    )


# ══════════════════════════════════════════════════════════════════════════
#  CALIFICACIONES — Admin
# ══════════════════════════════════════════════════════════════════════════

@app.route("/admin/calificaciones")
@requiere_login
@requiere_admin
def admin_calificaciones():
    """Panel de todas las calificaciones recibidas con estadísticas."""
    u = usuario_logueado()

    calificaciones = (
        db.query(Calificacion)
        .join(Calificacion.reserva)
        .order_by(Calificacion.fecha.desc())
        .all()
    )

    # Promedio global
    promedio_global = None
    if calificaciones:
        promedio_global = round(sum(c.puntuacion for c in calificaciones) / len(calificaciones), 1)

    # Promedios por espacio
    promedios_espacio = {}
    for cal in calificaciones:
        for esp in cal.reserva.espacios:
            if esp.nombre not in promedios_espacio:
                promedios_espacio[esp.nombre] = []
            promedios_espacio[esp.nombre].append(cal.puntuacion)
    promedios_espacio = {
        k: round(sum(v) / len(v), 1)
        for k, v in promedios_espacio.items()
    }

    # Distribución de estrellas
    dist_estrellas = {i: 0 for i in range(1, 6)}
    for cal in calificaciones:
        dist_estrellas[cal.puntuacion] = dist_estrellas.get(cal.puntuacion, 0) + 1

    return render_template(
        "admin_calificaciones.html",
        usuario=u,
        calificaciones=calificaciones,
        promedio_global=promedio_global,
        promedios_espacio=promedios_espacio,
        dist_estrellas=dist_estrellas,
        total_cals=len(calificaciones),
    )


@app.route("/admin/calificaciones/<int:cid>/responder", methods=["POST"])
@requiere_login
@requiere_admin
def admin_responder_calificacion(cid):
    """El admin responde a la calificación de un usuario."""
    cal = db.get(Calificacion, cid)
    if not cal:
        flash("Calificación no encontrada.", "danger")
        return redirect(url_for("admin_calificaciones"))

    respuesta = request.form.get("respuesta", "").strip()
    if not respuesta:
        flash("La respuesta no puede estar vacía.", "warning")
        return redirect(url_for("admin_calificaciones"))

    try:
        cal.respuesta_admin = respuesta
        db.commit()
        flash("Respuesta enviada correctamente.", "success")
    except Exception:
        db.rollback()
        flash("Error al guardar la respuesta.", "danger")

    return redirect(url_for("admin_calificaciones"))


if __name__ == "__main__":
    from database import crear_base_de_datos
    crear_base_de_datos()
    app.run(debug=True)
